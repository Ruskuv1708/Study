from datetime import date, datetime, time, timedelta, timezone
from tempfile import SpooledTemporaryFile

from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy.orm import Session

from core.config import settings

# Import the models we want to report on
from modules.workflow.workflow_models import Request
from modules.access_control.access_models import User
from modules.access_control.access_enums import UserRole

class ReportService:

    @staticmethod
    def _resolve_period(date_from: date | None, date_to: date | None):
        start_dt = None
        end_dt = None
        if date_from:
            start_dt = datetime.combine(date_from, time.min, tzinfo=timezone.utc)
        if date_to:
            end_dt = datetime.combine(date_to + timedelta(days=1), time.min, tzinfo=timezone.utc)
        return start_dt, end_dt

    @staticmethod
    def generate_requests_excel(
        db: Session,
        workspace_id,
        date_from: date | None = None,
        date_to: date | None = None,
    ):
        """
        Generates an Excel file containing all Requests for this company.
        """
        start_dt, end_dt = ReportService._resolve_period(date_from, date_to)
        # 1. Fetch Data (SQLAlchemy)
        # We join with User to get the actual name of the assignee
        query = db.query(
            Request.title,
            Request.status,
            Request.priority,
            Request.created_at,
            User.full_name.label("assignee_name")
        ).outerjoin(User, Request.assigned_to_id == User.id)\
         .filter(Request.workspace_id == workspace_id)
        if start_dt:
            query = query.filter(Request.created_at >= start_dt)
        if end_dt:
            query = query.filter(Request.created_at < end_dt)

        total = query.count()
        if total > settings.MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=f"Export exceeds max row limit ({settings.MAX_EXPORT_ROWS})"
            )
        # 2. Stream rows into a write-only workbook to avoid large memory spikes
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Requests_Export")
        sheet.append(["Title", "Status", "Priority", "Created Date", "Assigned To"])

        for row in query.yield_per(1000):
            sheet.append([
                row.title,
                row.status.value,
                row.priority.value,
                row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
                row.assignee_name if row.assignee_name else "Unassigned",
            ])

        output = SpooledTemporaryFile(max_size=settings.EXPORT_SPOOL_MAX_MB * 1024 * 1024)
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_users_excel(
        db: Session,
        workspace_id,
        date_from: date | None = None,
        date_to: date | None = None,
    ):
        """
        Another example: Export List of Employees
        """
        start_dt, end_dt = ReportService._resolve_period(date_from, date_to)
        user_query = db.query(User).filter(User.workspace_id == workspace_id)
        if start_dt:
            user_query = user_query.filter(User.created_at >= start_dt)
        if end_dt:
            user_query = user_query.filter(User.created_at < end_dt)
        total = user_query.count()
        if total > settings.MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=f"Export exceeds max row limit ({settings.MAX_EXPORT_ROWS})"
            )
        admin_roles = {UserRole.SUPERADMIN, UserRole.SYSTEM_ADMIN, UserRole.ADMIN}

        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Employees")
        sheet.append(["Name", "Email", "Role", "Admin"])
        for user in user_query.yield_per(1000):
            role_value = user.role.value if hasattr(user.role, "value") else str(user.role)
            is_admin = (
                user.role in admin_roles
                if hasattr(user.role, "value")
                else str(user.role).upper() in {r.value for r in admin_roles}
            )
            sheet.append([user.full_name, user.email, role_value, is_admin])

        output = SpooledTemporaryFile(max_size=settings.EXPORT_SPOOL_MAX_MB * 1024 * 1024)
        workbook.save(output)
        output.seek(0)
        return output
